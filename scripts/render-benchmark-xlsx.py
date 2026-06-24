#!/usr/bin/env python3
"""Benchmark Excel: two sections (veth + WAN), 5-run averages, embedded ba
charts. All chart data lives on the same sheet as the charts so series bind
correctly (avoids zero-series invalid charts).
"""
import argparse
import json
import statistics
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.chart import BarChart, Reference

IMPLS = [("rust", "this port"), ("go", "Go upstream"), ("jrs", "other rs")]
PERF_METRICS = [
    ("tcp_connect_rtt", "median_ms"),
    ("tcp_persistent_rtt", "median_ms"),
    ("tcp_throughput_mbps", "median"),
    ("udp_rtt", "median_ms"),
]


def mean_round(xs, nd=3):
    xs = [x for x in xs if x is not None]
    return round(statistics.mean(xs), nd) if xs else None


def load_perf(path):
    rows = [json.loads(l) for l in Path(path).read_text().splitlines() if l.strip()]
    out = {}
    for tag, _ in IMPLS:
        out[tag] = {}
        for metric, field in PERF_METRICS:
            vals = [float(r[metric][field]) for r in rows
                    if r.get("tag") == tag and r.get(metric) and r[metric].get(field) is not None]
            out[tag][metric] = mean_round(vals)
    return out


def load_mem(path):
    rows = [json.loads(l) for l in Path(path).read_text().splitlines() if l.strip()]
    out = {}
    for tag, _ in IMPLS:
        peaks = [float(r["server_peak_rss_kb"]) for r in rows
                 if r.get("tag") == tag and r.get("server_peak_rss_kb")]
        out[tag] = {"server_peak_rss_mb": round(statistics.mean(peaks) / 1024.0, 1) if peaks else None}
    return out


HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=11)
SECT_FILL = PatternFill("solid", fgColor="2E5496")
SECT_FONT = Font(color="FFFFFF", bold=True, size=12)
TITLE_FONT = Font(bold=True, size=15)
NOTE_FONT = Font(italic=True, size=9, color="666666")
THIN = Side(style="thin", color="B7C3D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
BEST_FILL = PatternFill("solid", fgColor="C6EFCE")
BEST_FONT = Font(bold=True, color="006100")
IMPL_FILLS = {
    "rust": PatternFill("solid", fgColor="E2EFDA"),
    "go": PatternFill("solid", fgColor="DDEBF7"),
    "jrs": PatternFill("solid", fgColor="FCE4D6"),
}
ROWS = [
    ("TCP connect+RTT (ms)", "tcp_connect_rtt", "perf", "lower"),
    ("TCP persistent RTT (ms)", "tcp_persistent_rtt", "perf", "lower"),
    ("TCP throughput (Mbps)", "tcp_throughput_mbps", "perf", "higher"),
    ("UDP RTT (ms)", "udp_rtt", "perf", "lower"),
    ("Server peak RSS (MB)", "server_peak_rss_mb", "mem", "lower"),
]


def write_table(ws, start_row, section_title, perf, mem, note):
    r = start_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    cell = ws.cell(row=r, column=1, value=section_title)
    cell.fill, cell.font, cell.alignment = SECT_FILL, SECT_FONT, LEFT
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(row=r, column=1, value=note).font = NOTE_FONT
    r += 1
    for ci, h in enumerate(["Metric", "this port", "Go upstream", "other rs", "Better"], start=1):
        c = ws.cell(row=r, column=ci, value=h)
        c.fill, c.font, c.alignment, c.border = HDR_FILL, HDR_FONT, CENTER, BORDER
    r += 1
    vals_by_key = {}
    for label, key, kind, direction in ROWS:
        ws.cell(row=r, column=1, value=label).alignment = LEFT
        ws.cell(row=r, column=1).border = BORDER
        vals = {tag: (perf[tag].get(key) if kind == "perf" else mem[tag].get(key)) for tag, _ in IMPLS}
        vals_by_key[key] = vals
        present = [v for v in vals.values() if v is not None]
        best = (min(present) if direction == "lower" else max(present)) if present else None
        for ci, (tag, _) in enumerate(IMPLS, start=2):
            v = vals[tag]
            c = ws.cell(row=r, column=ci, value=v)
            c.alignment, c.border, c.fill = RIGHT, BORDER, IMPL_FILLS[tag]
            if isinstance(v, float):
                c.number_format = "0.000" if v < 100 else "0.0"
            if best is not None and v is not None and abs(v - best) < 1e-9:
                c.fill, c.font = BEST_FILL, BEST_FONT
        b = ws.cell(row=r, column=5, value=direction)
        b.alignment, b.border = CENTER, BORDER
        r += 1
    return r, vals_by_key


def build_chart(ws, data_top, n_metrics, title):
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.height = 7.0
    chart.width = 16
    data = Reference(ws, min_col=2, max_col=4, min_row=data_top, max_row=data_top + n_metrics)
    cats = Reference(ws, min_col=1, min_row=data_top + 1, max_row=data_top + n_metrics)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    # Excel rejects a column chart whose category axis is positioned left;
    # openpyxl defaults both axes to left, so set them explicitly.
    chart.x_axis.axPos = "b"
    chart.y_axis.axPos = "l"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    return chart


def add_chart(ws, chart, anchor):
    ws.add_chart(chart, anchor)


def write_data_block(ws, top, header_label, metric_rows):
    ws.cell(row=top, column=1, value=header_label)
    for ci, (_, lab) in enumerate(IMPLS, start=2):
        ws.cell(row=top, column=ci, value=lab)
    rr = top + 1
    for label, vals in metric_rows:
        ws.cell(row=rr, column=1, value=label)
        for ci, (tag, _) in enumerate(IMPLS, start=2):
            ws.cell(row=rr, column=ci, value=vals.get(tag))
        rr += 1
    chart = build_chart(ws, top, len(metric_rows), header_label)
    return chart, rr - 1


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--veth-perf", required=True)
    ap.add_argument("--veth-mem", required=True)
    ap.add_argument("--wan-perf", required=True)
    ap.add_argument("--wan-mem", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    vperf, vmem = load_perf(args.veth_perf), load_mem(args.veth_mem)
    wperf, wmem = load_perf(args.wan_perf), load_mem(args.wan_mem)

    wb = Workbook()
    ws = wb.active
    ws.title = "summary"

    ws.merge_cells("A1:E1")
    ws.cell(row=1, column=1, value="zuicity benchmark").font = TITLE_FONT
    ws.merge_cells("A2:E2")
    ws.cell(row=2, column=1,
            value="Average of 5 runs per implementation. Best value per row highlighted green.").font = NOTE_FONT

    after1, v_rows = write_table(
        ws, 4, "Local (two network namespaces, veth)", vperf, vmem,
        "Two Linux network namespaces joined by a veth pair; real QUIC, not loopback.")
    after2, w_rows = write_table(
        ws, after1 + 2, "WAN (Oracle server-to-server, real Internet)", wperf, wmem,
        "Servers on Oracle host A (134.185.121.128); clients on Oracle host B (158.180.72.138).")

    # Charts sit to the right of the data tables in two non-overlapping columns.
    # Each chart is 9.5cm wide (~7 default-width columns), so the two chart
    # stacks must be at least that far apart horizontally. LOCAL charts anchor
    # at column G (col 7); WAN charts anchor at column Q (col 17), leaving a
    # clear gap so the two columns never collide. Charts step by 11 rows.
    veth_start = 4
    wan_start = after1 + 2
    block = after2 + 3
    for sect, rows, sect_top, chart_col in (("veth (local)", v_rows, veth_start, "G"),
                                            ("WAN", w_rows, wan_start, "Q")):
        specs = [
            (f"{sect} latency", f"{sect}: Latency (ms, lower=better)",
             [("connect+RTT", rows["tcp_connect_rtt"]),
              ("persistent RTT", rows["tcp_persistent_rtt"]),
              ("UDP RTT", rows["udp_rtt"])]),
            (f"{sect} throughput", f"{sect}: Throughput (Mbps, higher=better)",
             [("throughput", rows["tcp_throughput_mbps"])]),
            (f"{sect} server RSS", f"{sect}: Server RSS (MB, lower=better)",
             [("server RSS", rows["server_peak_rss_mb"])]),
        ]
        anchor_row = sect_top
        for header, title, metric_rows in specs:
            chart, next_block = write_data_block(ws, block, header, metric_rows)
            chart.title = title
            chart.height = 5.2
            chart.width = 9.5
            add_chart(ws, chart, f"{chart_col}{anchor_row}")
            block = next_block + 2
            anchor_row += 11

    ws.column_dimensions["A"].width = 26
    for ci in ("B", "C", "D"):
        ws.column_dimensions[ci].width = 16
    ws.column_dimensions["E"].width = 9

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()